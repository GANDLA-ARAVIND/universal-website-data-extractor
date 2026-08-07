import io
from typing import Any, List, Optional, Tuple, Union

import openpyxl  # type: ignore
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

from src.application.services.exporters.base import BaseExporter
from src.db.models.crawl_job import CrawlJob
from src.db.models.page import ExtractedPage
from src.db.models.statistic import CrawlStatistic
from src.schemas.export import ExportFormat


def clean_val(val: Any) -> str:
    """Safely cleans strings for openpyxl cell output by removing illegal control characters."""
    if val is None:
        return ""
    s = str(val)
    return "".join(ch for ch in s if ord(ch) >= 32 or ch in "\n\r\t")


class XlsxExporter(BaseExporter):
    """Strategy Exporter for multi-tab Microsoft Excel (.xlsx) workbook generation."""

    @property
    def format_type(self) -> ExportFormat:
        return ExportFormat.XLSX

    async def export(
        self,
        pages: List[ExtractedPage],
        job: CrawlJob,
        stats: Optional[CrawlStatistic] = None,
    ) -> Tuple[bytes, str, str]:
        wb = openpyxl.Workbook()

        # Styling Definitions
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        def style_headers(ws: Any, headers: List[str]) -> None:
            ws.append([clean_val(h) for h in headers])
            ws.freeze_panes = "A2"
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        def auto_fit_columns(ws: Any) -> None:
            for col in ws.columns:
                max_len = 0
                col_cells = list(col)
                if not col_cells:
                    continue
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

        # =====================================================================
        # 1. WORKSHEET: OVERVIEW
        # =====================================================================
        ws_overview: Any = wb.active
        ws_overview.title = "Overview"

        # Enable grid lines safely
        if hasattr(ws_overview, "sheet_view"):
            ws_overview.sheet_view.showGridLines = True

        ws_overview.append(["Website Crawl Executive Overview"])
        ws_overview.cell(row=1, column=1).font = title_font
        ws_overview.append([])

        duration_val: float = (
            float(stats.total_duration_sec)
            if stats is not None and getattr(stats, "total_duration_sec", None) is not None
            else 0.0
        )
        total_links: int = (
            int(stats.total_links)
            if stats is not None and getattr(stats, "total_links", None) is not None
            else sum(len(getattr(p, "links", []) or []) for p in pages)
        )
        total_images: int = (
            int(stats.total_images)
            if stats is not None and getattr(stats, "total_images", None) is not None
            else sum(len(getattr(p, "images", []) or []) for p in pages)
        )

        job_status_obj = getattr(job, "status", "PENDING")
        job_status_str = (
            job_status_obj.value if hasattr(job_status_obj, "value") else str(job_status_obj)
        )

        job_created = getattr(job, "created_at", None)
        created_str = (
            job_created.strftime("%Y-%m-%d %H:%M:%S UTC")
            if job_created is not None
            else "N/A"
        )

        overview_data: List[Tuple[str, Union[str, int, float, bool]]] = [
            ("Job ID", str(job.id)),
            ("Target Seed URL", clean_val(job.seed_url)),
            ("Execution Status", clean_val(job_status_str)),
            ("Max Depth Setting", job.max_depth),
            ("Max Pages Setting", job.max_pages),
            ("Render JS Flag", job.render_js),
            ("Extracted Pages Total", len(pages)),
            ("Discovered Links Total", total_links),
            ("Discovered Images Total", total_images),
            ("Total Duration (seconds)", duration_val),
            ("Created Timestamp", created_str),
        ]

        for label, val in overview_data:
            current_row: int = int(ws_overview.max_row or 0) + 1
            c1 = ws_overview.cell(row=current_row, column=1, value=clean_val(label))
            c1.font = bold_font
            c1.border = thin_border

            c2 = ws_overview.cell(
                row=current_row,
                column=2,
                value=clean_val(val) if isinstance(val, str) else val,
            )
            c2.font = regular_font
            c2.border = thin_border

        auto_fit_columns(ws_overview)

        # =====================================================================
        # 2. WORKSHEET: PAGES
        # =====================================================================
        ws_pages: Any = wb.create_sheet(title="Pages")
        page_headers = [
            "URL",
            "Normalized URL",
            "Status Code",
            "Depth",
            "Title",
            "Meta Description",
            "Links Count",
            "Images Count",
            "Response Time (ms)",
            "Fetched Timestamp",
        ]
        style_headers(ws_pages, page_headers)

        for p in pages:
            p_created = getattr(p, "created_at", None)
            page_created_str = p_created.isoformat() if p_created is not None else ""
            p_links = getattr(p, "links", []) or []
            p_images = getattr(p, "images", []) or []

            ws_pages.append(
                [
                    clean_val(p.url),
                    clean_val(p.normalized_url),
                    p.status_code,
                    p.depth,
                    clean_val(p.title),
                    clean_val(p.meta_description),
                    len(p_links),
                    len(p_images),
                    p.response_time_ms,
                    page_created_str,
                ]
            )

        max_p_row: int = int(ws_pages.max_row or 1)
        if max_p_row > 1:
            ws_pages.auto_filter.ref = f"A1:J{max_p_row}"
        auto_fit_columns(ws_pages)

        # =====================================================================
        # 3. WORKSHEET: LINKS
        # =====================================================================
        ws_links: Any = wb.create_sheet(title="Links")
        link_headers = [
            "Source Page URL",
            "Discovered Target URL",
            "Anchor Text",
            "Is External",
        ]
        style_headers(ws_links, link_headers)

        for p in pages:
            p_links = getattr(p, "links", []) or []
            for l in p_links:
                ws_links.append(
                    [
                        clean_val(l.source_url),
                        clean_val(l.target_url),
                        clean_val(l.anchor_text),
                        l.is_external,
                    ]
                )

        max_l_row: int = int(ws_links.max_row or 1)
        if max_l_row > 1:
            ws_links.auto_filter.ref = f"A1:D{max_l_row}"
        auto_fit_columns(ws_links)

        # =====================================================================
        # 4. WORKSHEET: IMAGES
        # =====================================================================
        ws_images: Any = wb.create_sheet(title="Images")
        img_headers = ["Page URL", "Image Source URL", "Alt Text"]
        style_headers(ws_images, img_headers)

        for p in pages:
            p_images = getattr(p, "images", []) or []
            for img in p_images:
                ws_images.append(
                    [
                        clean_val(p.url),
                        clean_val(img.image_url),
                        clean_val(img.alt_text),
                    ]
                )

        max_img_row: int = int(ws_images.max_row or 1)
        if max_img_row > 1:
            ws_images.auto_filter.ref = f"A1:C{max_img_row}"
        auto_fit_columns(ws_images)

        # =====================================================================
        # 5. WORKSHEET: STATISTICS
        # =====================================================================
        ws_stats: Any = wb.create_sheet(title="Statistics")
        stat_headers = ["Metric Name", "Metric Value"]
        style_headers(ws_stats, stat_headers)

        crawled_cnt: int = (
            int(stats.pages_crawled)
            if stats is not None and getattr(stats, "pages_crawled", None) is not None
            else len(pages)
        )
        failed_cnt: int = (
            int(stats.failed_pages)
            if stats is not None and getattr(stats, "failed_pages", None) is not None
            else 0
        )

        stats_rows: List[Tuple[str, Union[int, float]]] = [
            ("Pages Crawled Count", crawled_cnt),
            ("Failed Pages Count", failed_cnt),
            ("Total Images Extracted", total_images),
            ("Total Links Discovered", total_links),
            ("Total Execution Duration (s)", duration_val),
        ]
        for name, val_num in stats_rows:
            ws_stats.append([clean_val(name), val_num])

        auto_fit_columns(ws_stats)

        # Build Bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()
        buffer.close()

        filename = self.generate_filename(job.seed_url, str(job.id), "xlsx")
        return (
            xlsx_bytes,
            filename,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
